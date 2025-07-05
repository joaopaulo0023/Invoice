import streamlit as st
import pdfplumber
import pandas as pd
import io
import re
import json
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment, Border, Side

# Título
st.title("SHIPMENT FOLLOW UP - TIME EXPEDIÇÃO")

# Arquivo para armazenar o histórico
HISTORY_FILE = "janelas_history.json"

def load_history():
    try:
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_history(history):
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

# Extrai janelas do PDF
def extract_windows_from_pdf(pdf_bytes):
    data = {"Client": [], "Time Window": [], "Soldto": []}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            lines = (page.extract_text() or "").split("\n")
            for i, line in enumerate(lines):
                if re.match(r"^(\d{2}:\d{2} \- \d{2}:\d{2})$", line.strip()):
                    tw = line.strip()
                    client = lines[i+1].strip() if i+1 < len(lines) else ""
                    soldto = ""
                    if i+2 < len(lines):
                        m = re.search(r"\((\d+)\)", lines[i+2])
                        soldto = m.group(1) if m else ""
                    data["Client"].append(client)
                    data["Time Window"].append(tw)
                    data["Soldto"].append(soldto)
    return pd.DataFrame(data)

# Carrega histórico salvo
history = load_history()

# Upload do PDF
uploaded_file = st.file_uploader("Faça upload do PDF de janelas de expedição", type=["pdf"])
if uploaded_file:
    with st.spinner("Extraindo janelas..."):
        df = extract_windows_from_pdf(uploaded_file.read())

    # ➕ Campo para adicionar janela manual (com expander)
    with st.expander("➕ Adicionar manualmente uma nova janela"):
        with st.form("manual_window_form"):
            client_manual = st.text_input("Cliente")
            time_window_manual = st.text_input("Janela de Tempo (ex: 08:00 - 10:00)")
            soldto_manual = st.text_input("Soldto (código numérico)")
            submitted = st.form_submit_button("Adicionar janela manualmente")

            if submitted:
                if client_manual and time_window_manual and soldto_manual:
                    new_row = {
                        "Client": client_manual.strip(),
                        "Time Window": time_window_manual.strip(),
                        "Soldto": soldto_manual.strip()
                    }
                    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                    st.success("Janela adicionada com sucesso!")
                else:
                    st.error("Preencha todos os campos antes de adicionar.")

    if df.empty:
        st.warning("Nenhuma janela de tempo encontrada no PDF.")
    else:
        st.success(f"Encontradas {len(df)} janelas de tempo")        
        st.markdown("---")
        st.subheader("Marcar Janelas e Definir Status")

        output_records = []
        for idx, row in df.iterrows():
            cols = st.columns([3, 2, 1, 1, 2, 3])
            cols[0].write(row['Client'])
            cols[1].write(row['Time Window'])
            cols[2].write(row['Soldto'])
            keep = cols[3].checkbox("Manter", value=True, key=f"keep_{idx}")
            if not keep:
                continue
            key = f"{row['Soldto']}_{row['Time Window']}"
            prev = history.get(key, {})
            status = cols[4].selectbox(
                "Status", ("CARREGADO", "Não saiu"),
                index=0 if prev.get('status') == 'CARREGADO' else 1,
                key=f"status_{idx}"
            )
            obs = cols[5].text_input(
                "Observação", value=prev.get('obs', ''),
                key=f"obs_{idx}"
            )
            history[key] = {'status': status, 'obs': obs, 'last_updated': datetime.now().isoformat()}
            output_records.append({
                'Client': row['Client'],
                'Time Window': row['Time Window'],
                'Soldto': row['Soldto'],
                'Status': status,
                'Observação': obs
            })
        save_history(history)

        final_df = pd.DataFrame(output_records)

        with st.expander("DataFrame Inicial e Final"):
            st.subheader("DataFrame Inicial")
            st.dataframe(df)
            st.subheader("Resumo Final")
            st.dataframe(final_df)

        # Gera Excel formatado
        output = io.BytesIO()
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            final_df.to_excel(writer, index=False, sheet_name="Janelas")
            ws = writer.sheets['Janelas']

            # Ajusta larguras de coluna
            widths = {'A': 25, 'B': 18, 'C': 12, 'D': 15, 'E': 50}
            for col, w in widths.items():
                ws.column_dimensions[col].width = w

            # Ajusta alinhamento e bordas
            max_row = len(final_df) + 1
            for row in range(1, max_row + 1):
                for col in ['A', 'B', 'C', 'D', 'E']:
                    cell = ws[f"{col}{row}"]
                    cell.alignment = Alignment(wrap_text=True)
                    cell.border = border

            # Congela primeira linha
            ws.freeze_panes = ws['A2']

            # Cores conforme status
            green = PatternFill(start_color='C6EFCE', fill_type='solid')
            red = PatternFill(start_color='FFC7CE', fill_type='solid')
            for row_idx in range(2, max_row + 1):
                status_val = ws.cell(row=row_idx, column=4).value
                fill = green if status_val == 'CARREGADO' else red if status_val == 'Não saiu' else None
                if fill:
                    for col_letter in ['A', 'B', 'C', 'D', 'E']:
                        ws[f"{col_letter}{row_idx}"].fill = fill

        processed_data = output.getvalue()
        st.download_button(
            "📥 Baixar relatório Excel formatado", processed_data,
            file_name="janelas_expedicao_formatado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
