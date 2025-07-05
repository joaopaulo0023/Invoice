import pandas as pd
import sqlite3
import os
from openpyxl import load_workbook

# Configurações iniciais
DB_NAME = "gestao_operacional.db"
EXCEL_FILE = r"S:\ABRCommon\Departamentos\AMC Logistica\03 - Operacoes\10 - Projetos&Melhorias\Quadro Gestão Visual MOD\Quadro Gestão Visual Mão de Obra - Logística.xlsx"
FOTOS_DIR = "fotos_operadores"  # Pasta para armazenar as imagens

# Criar diretório para fotos (se não existir)
os.makedirs(FOTOS_DIR, exist_ok=True)

# Conectar ao banco de dados
conn = sqlite3.connect(DB_NAME)
cursor = conn.cursor()

# Criar tabelas (execute apenas uma vez)
cursor.execute('''
    CREATE TABLE IF NOT EXISTS Turnos (
        TurnoID INTEGER PRIMARY KEY,
        NomeTurno TEXT NOT NULL
    )
''')

cursor.execute('''
    CREATE TABLE IF NOT EXISTS Operadores (
        OperadorID INTEGER PRIMARY KEY,
        Nome TEXT NOT NULL,
        Cargo TEXT,
        TurnoID INTEGER,
        Foto TEXT,
        FOREIGN KEY (TurnoID) REFERENCES Turnos(TurnoID)
    )
''')

cursor.execute('''
    CREATE TABLE IF NOT EXISTS Metricas (
        MetricaID INTEGER PRIMARY KEY,
        TipoMetrica TEXT NOT NULL,
        Valor REAL,
        Data DATE,
        OperadorID INTEGER,
        FOREIGN KEY (OperadorID) REFERENCES Operadores(OperadorID)
    )
''')

# Função para processar a planilha de operadores (exemplo)
def importar_operadores():
    wb = load_workbook(EXCEL_FILE)
    sheet = wb["DISTR. OPERACIONAL"]  # Altere para o nome correto da sheet
    
    # Exemplo de leitura de dados - ajuste os índices conforme seu arquivo
    for row in sheet.iter_rows(min_row=7, max_col=4, values_only=True):
        nome, cargo, turno, foto_path = row  # Adapte às suas colunas
        
        # Inserir turno se não existir
        cursor.execute('''
            INSERT OR IGNORE INTO Turnos (NomeTurno) 
            VALUES (?)
        ''', (turno,))
        
        # Obter ID do turno
        cursor.execute('SELECT TurnoID FROM Turnos WHERE NomeTurno = ?', (turno,))
        turno_id = cursor.fetchone()[0]
        
        # Inserir operador
        cursor.execute('''
            INSERT INTO Operadores (Nome, Cargo, TurnoID, Foto)
            VALUES (?, ?, ?, ?)
        ''', (nome, cargo, turno_id, foto_path))

# Função para processar métricas (exemplo)
def importar_metricas():
    wb = load_workbook(EXCEL_FILE)
    sheet = wb["GESTÃO LIDERANÇA"]  # Altere para a sheet correta
    
    # Exemplo de leitura - ajuste conforme necessidade
    for row in sheet.iter_rows(min_row=10, values_only=True):
        operador_nome, metrica_tipo, valor, data = row  # Adaptar
        
        # Obter ID do operador
        cursor.execute('SELECT OperadorID FROM Operadores WHERE Nome = ?', (operador_nome,))
        operador_id = cursor.fetchone()[0]
        
        # Inserir métrica
        cursor.execute('''
            INSERT INTO Metricas (TipoMetrica, Valor, Data, OperadorID)
            VALUES (?, ?, ?, ?)
        ''', (metrica_tipo, valor, data, operador_id))

# Executar importação
try:
    importar_operadores()
    importar_metricas()
    conn.commit()
    print("Dados importados com sucesso!")
except Exception as e:
    conn.rollback()
    print(f"Erro: {str(e)}")
finally:
    conn.close()