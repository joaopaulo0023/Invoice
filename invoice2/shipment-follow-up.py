import streamlit as st
import pdfplumber
import pandas as pd
import io
import re
import json
from datetime import datetime, date, time, timedelta
from openpyxl.styles import PatternFill, Alignment, Border, Side

st.title("SHIPMENT FOLLOW UP - TIME EXPEDIÇÃO")

HISTORY_FILE = "janelas_history.json"
LAST_DAY_FILE = "last_day.txt"

# Zera o JSON apenas após 06:00 do novo dia
def reset_history_if_new_day():
    agora = datetime.now()
    corte = time(6, 0)  # 06:00
    # Se antes das 06:00, consideramos "ontem" como boundary_date
    if agora.time() < corte:
        boundary_date = (agora.date() - timedelta(days=1)).isoformat()
    else:
        boundary_date = agora.date().isoformat()

    try:
        with open(LAST_DAY_FILE, "r") as f:
            last_day = f.read().strip()
    except FileNotFoundError:
        last_day = ""

    # Só reseta quando boundary_date mudar
    if last_day != boundary_date:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump({}, f)
        with open(LAST_DAY_FILE, "w") as f:
            f.write(boundary_date)

reset_history_if_new_day()

def load_history():
    try:
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

# Carrega histórico e garante a chave de hoje
history = load_history()

# Data "hoje" usada apenas para chaves em memória; não afeta reset
hoje = date.today().isoformat()
history.setdefault(hoje, {})

turnos = {
    "1º Turno (06:00–15:10)": ("06:00", "15:10"),
    "2º Turno (15:10–23:10)": ("15:10", "23:10"),
    "3º Turno (23:10–06:00)": ("23:10", "06:00"),
}
for t in turnos:
    history[hoje].setdefault(t, {})

# Determina turno atual
def determinar_turno_atual():
    agora = datetime.now().time()
    h0600 = time(6, 0)
    h1510 = time(15, 10)
    h2310 = time(23, 10)
    if h0600 <= agora < h1510:
        return "1º Turno (06:00–15:10)"
    elif h1510 <= agora < h2310:
        return "2º Turno (15:10–23:10)"
    else:
        return "3º Turno (23:10–06:00)"

selected_turno = determinar_turno_atual()
st.success(f"Turno detectado automaticamente: **{selected_turno}**")
turn_history = history[hoje][selected_turno]

# Extrai janelas do PDF
def extract_windows_from_pdf(pdf_bytes):
    data = {"Client": [], "Time Window": [], "Soldto": []}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            lines = (page.extract_text() or "").split("\n")
            for i, line in enumerate(lines):
                if re.match(r"^\d{2}:\d{2} \- \d{2}:\d{2}$", line.strip()):
                    tw = line.strip()
                    client = lines[i+1].strip() if i+1 < len(lines) else ""
                    soldto = ""
                    if i+2 < len(lines):
                        m = re.search(r"\((\d+)\)", lines[i+2])
                        soldto = m.group(1) if m else ""
                    data["Client"].append(client)
                    data["Time Window"].append(tw)
                    data["Soldto"].append(soldto)
    return pd.DataFrame(data)

# Sessão: PDF enviado

if "uploaded_pdf_df" not in st.session_state:
    st.session_state.uploaded_pdf_df = pd.DataFrame()
    st.session_state.uploaded_name = None

uploaded_file = st.file_uploader("Upload do PDF de janelas", type=["pdf"])
if uploaded_file and uploaded_file.name != st.session_state.uploaded_name:
    with st.spinner("Extraindo janelas..."):
        st.session_state.uploaded_pdf_df = extract_windows_from_pdf(uploaded_file.read())
    st.session_state.uploaded_name = uploaded_file.name

if st.session_state.uploaded_pdf_df.empty:
    st.info("Faça o upload do PDF de janelas para começar a visualizar e marcar as janelas.")
else:
    pdf_df = st.session_state.uploaded_pdf_df.copy()
    manual_records = []
    for turno_hist in history[hoje].values():
        for entry in turno_hist.values():
            if entry.get("client") and entry.get("time_window") and entry.get("soldto"):
                manual_records.append({
                    "Client": entry["client"],
                    "Time Window": entry["time_window"],
                    "Soldto": entry["soldto"]
                })
    manual_df = pd.DataFrame(manual_records)
    combined = pd.concat([pdf_df, manual_df], ignore_index=True) if not pdf_df.empty or not manual_df.empty else pd.DataFrame(columns=["Client","Time Window","Soldto"])
    combined.drop_duplicates(subset=["Soldto", "Time Window"], inplace=True)
    st.session_state.df = combined

    with st.expander("➕ Adicionar janela manual"):
        with st.form("manual_form", clear_on_submit=True):
            c = st.text_input("Cliente", key="client_manual")
            tw = st.text_input("Janela de Tempo (ex: 08:00 - 10:00)", key="time_window_manual")
            s = st.text_input("Soldto (código numérico)", key="soldto_manual")
            if st.form_submit_button("Adicionar"):
                if c and tw and s:
                    key = f"{s.strip()}_{tw.strip()}"
                    for t in turnos:
                        history[hoje][t][key] = {**history[hoje][t].get(key, {}),
                            "client": c.strip(),
                            "time_window": tw.strip(),
                            "soldto": s.strip(),
                            "status": "",
                            "obs": "",
                            "exit_time": "",
                            "pallets": 0,
                            "last_updated": datetime.now().isoformat()
                        }
                    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
                        json.dump(history, f, ensure_ascii=False, indent=2)
                    st.success("Janela manual adicionada!")
                else:
                    st.error("Preencha todos os campos.")

    st.markdown("---")
    st.subheader("Marcar Janelas, Paletes Carregados e Observações")
    output = []
    for idx, row in st.session_state.df.iterrows():
        cols = st.columns([3,2,1,1,2,2,3,2])
        cols[0].write(row["Client"])
        cols[1].write(row["Time Window"])
        cols[2].write(row["Soldto"])
        if not cols[3].checkbox("Manter", True, key=f"keep_{idx}"):
            continue
        key = f"{row['Soldto']}_{row['Time Window']}"
        turn_history.setdefault(key, {"pallets":0, "status":"","obs":"","exit_time":"","last_updated":""})
        prev = turn_history[key]
        status = cols[4].selectbox("Status", ["CARREGADO","Não saiu"], index=0 if prev.get("status")=="CARREGADO" else 1, key=f"status_{idx}")
        pallets = cols[5].number_input("Paletes Carregados", min_value=0, step=1, value=prev.get("pallets",0), key=f"pallets_{idx}")
        obs = cols[6].text_area("Observação", value=prev.get("obs",""), key=f"obs_{idx}", height=80)
        exit_time = cols[7].text_input("Hora Saída", value=prev.get("exit_time",""), key=f"exit_{idx}")
        turn_history[key].update({"status":status,"obs":obs,"exit_time":exit_time,"pallets":pallets,"last_updated":datetime.now().isoformat()})
        output.append({"Client":row["Client"],"Time Window":row["Time Window"],"Soldto":row["Soldto"],"Status":status,"Paletes Carregados":pallets,"Observação":obs,"Hora Saída":exit_time})

    # Salva histórico após interações
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

    total_paletes_turno = sum(item["Paletes Carregados"] for item in output)
    total_paletes_dia = sum(entry.get("pallets",0) for turno_hist in history[hoje].values() for entry in turno_hist.values())
    final_df = pd.DataFrame(output)

    # Determina turno de saída por linha
    final_df["Turno"] = ""
    for i, row in final_df.iterrows():
        try:
            if row["Hora Saída"]:
                saida = datetime.strptime(row["Hora Saída"], "%H:%M").time()
                for turno, (inicio, fim) in turnos.items():
                    ini = datetime.strptime(inicio, "%H:%M").time()
                    fim = datetime.strptime(fim, "%H:%M").time()
                    if ini < fim:
                        if ini <= saida < fim:
                            final_df.at[i, "Turno"] = turno
                    else:
                        if saida >= ini or saida < fim:
                            final_df.at[i, "Turno"] = turno
        except:
            pass

    resumo_turnos = final_df.groupby("Turno")["Paletes Carregados"].sum().reset_index()
    col1, col2 = st.columns(2)
    col1.metric("Total de paletes no turno selecionado", total_paletes_turno)
    col2.metric("Total de paletes no dia", total_paletes_dia)

    # Exporta Excel formatado
    output_io = io.BytesIO()
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin,left=thin,right=thin,bottom=thin)
    with pd.ExcelWriter(output_io, engine="openpyxl") as writer:
        final_df.to_excel(writer, index=False, sheet_name="Janelas")
        resumo_turnos.to_excel(writer, index=False, sheet_name="ResumoTurnos")
        ws = writer.sheets["Janelas"]
        widths = {'A':25,'B':18,'C':12,'D':15,'E':20,'F':15,'G':50,'H':20,'I':25}
        for col,w in widths.items(): ws.column_dimensions[col].width = w
        max_row = len(final_df) + 1
        for r in range(1, max_row+1):
            for col in widths:
                cell = ws[f"{col}{r}"]
                cell.alignment = Alignment(wrap_text=True)
                cell.border = border
        ws.freeze_panes = ws["A2"]
        green = PatternFill(start_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", fill_type="solid")
        for r in range(2, max_row+1):
            fill = green if ws.cell(r,4).value == "CARREGADO" else red
            for col in widths: ws[f"{col}{r}"].fill = fill

        summary_row = max_row + 2
        turno_labels = {
            "1º Turno (06:00–15:10)": "Quantidade do 1ºT",
            "2º Turno (15:10–23:10)": "Quantidade do 2ºT",
            "3º Turno (23:10–06:00)": "Quantidade do 3ºT"
        }
        offset = 0
        for i, row in resumo_turnos.iterrows():
            label = turno_labels.get(row["Turno"], row["Turno"])
            ws[f"A{summary_row + offset}"] = label
            ws[f"B{summary_row + offset}"] = row["Paletes Carregados"]
            ws[f"A{summary_row + offset}"].font = ws[f"A{summary_row + offset}"].font.copy(bold=True)
            ws[f"B{summary_row + offset}"].font = ws[f"B{summary_row + offset}"].font.copy(bold=True)
            offset += 1

        ws[f"A{summary_row + offset}"] = "Total de paletes no dia"
        ws[f"B{summary_row + offset}"] = total_paletes_dia
        ws[f"A{summary_row + offset}"].font = ws[f"A{summary_row + offset}"].font.copy(bold=True)
        ws[f"B{summary_row + offset}"].font = ws[f"B{summary_row + offset}"].font.copy(bold=True)

    data_bytes = output_io.getvalue()
    st.download_button("📅 Baixar relatório Excel formatado", data=data_bytes, file_name="janelas_expedicao_formatado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
